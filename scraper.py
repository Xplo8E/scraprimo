from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment

def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36")
    chrome_options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
    service = Service("chromedriver-mac-arm64/chromedriver")  # Replace with your path
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def extract_nested_quests(soup):
    nested_quests = []
    nested_quest_links = soup.select("div.archive-style-wrapper > p > a")
    
    for link in nested_quest_links:
        quest_name = link.text.strip()
        quest_url = urljoin(base_url, link['href'])
        nested_quests.append((quest_name, quest_url))
    
    return nested_quests

def create_excel_file():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Genshin Impact World Quests"
    headers = ["Region", "Quest Name", "Rewards"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    return workbook, sheet

def write_to_excel(sheet, region, quest_name, rewards, row, is_sub_quest=False):
    if is_sub_quest:
        sheet.cell(row=row, column=2, value=quest_name)
    else:
        sheet.cell(row=row, column=1, value=region)
        sheet.cell(row=row, column=2, value=quest_name)
    
    rewards_cell = sheet.cell(row=row, column=3)
    rewards_text = "\n".join(rewards)
    rewards_cell.value = rewards_text
    rewards_cell.alignment = Alignment(wrapText=True)

def scrape_website(url, container_selector):
    driver = setup_driver()
    workbook, sheet = create_excel_file()
    current_row = 2  # Start from row 2 (after headers)

    print(f"Fetching {url}")
    response = requests.get(url)
    print("Loading content into Selenium")
    driver.execute_script(f"document.body.innerHTML = `{response.text}`")
    
    print("Searching for the quest information container...")
    try:
        container = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, container_selector))
        )
        print("Quest information container found.")
        
        # Find all tables with world quests
        tables = container.find_elements(By.CSS_SELECTOR, "table.a-table")
        print(f"Found {len(tables)} tables with world quests")
        
        for table_index, table in enumerate(tables, start=1):
            region_name = table.find_element(By.TAG_NAME, "th").text.strip()
            print(f"\nProcessing table {table_index}")
            print(f"Region: {region_name}")
            
            quest_links = table.find_elements(By.CSS_SELECTOR, "tbody > tr > td:nth-child(1) > a")
            print(f"Found {len(quest_links)} quest links")
            
            for i, link in enumerate(quest_links, 1):
                link_url = urljoin(url, link.get_attribute('href'))
                if not link_url.startswith('http'):
                    print(f"Skipping invalid URL: {link_url}")
                    continue
                print(f"\nVisiting link {i}/{len(quest_links)} for {region_name}: {link_url}")
                try:
                    response = requests.get(link_url)
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    # Extract quest name (title)
                    title = soup.title.string if soup.title else "No title found"
                    title = title.split('|')[0].strip()  # Remove the "| Genshin Impact｜Game8" part

                    print(f"Quest Name: {title}")

                    # Check for rewards first
                    rewards = process_rewards(soup, title)
                    
                    if rewards:
                        write_to_excel(sheet, region_name, title, rewards, current_row)
                        current_row += 1
                    else:
                        nested_quests = extract_nested_quests(soup)
                        if nested_quests:
                            write_to_excel(sheet, region_name, title, ["Chained Quest"], current_row)
                            current_row += 1
                            for nested_quest_name, nested_quest_url in nested_quests:
                                nested_response = requests.get(nested_quest_url)
                                nested_soup = BeautifulSoup(nested_response.text, 'html.parser')
                                nested_rewards = process_rewards(nested_soup, nested_quest_name)
                                if nested_rewards:
                                    write_to_excel(sheet, "", nested_quest_name, nested_rewards, current_row, is_sub_quest=True)
                                    current_row += 1
                        else:
                            print("No rewards or nested quests found")

                except Exception as e:
                    print(f"Error loading page or extracting info: {str(e)}")
                    print(f"Error occurred at line: {e.__traceback__.tb_lineno}")
                    continue
                
                time.sleep(1)  # Add a small delay between requests
            
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        driver.save_screenshot("error_screenshot.png")
        print("Screenshot saved as error_screenshot.png")
    
    driver.quit()
    workbook.save("genshin_impact_world_quests_rewards.xlsx")
    print("Results saved to genshin_impact_world_quests_rewards.xlsx")

def process_rewards(soup, quest_name):
    # Try the original selector first
    rewards_tables = soup.find_all('table', class_=['a-table', 'top', 'center'])

    # If no tables found, try the new selector
    if not rewards_tables:
        new_selector = "div.archive-style-wrapper > table.a-table.a-table.top.center"
        rewards_tables = soup.select(new_selector)

    if rewards_tables:
        print(f"Found {len(rewards_tables)} potential rewards tables")
    else:
        print("No rewards table found")
        return []

    rewards = []
    for reward_table in rewards_tables:
        # Check if the table contains 'reward' in its title or content
        table_text = reward_table.text.lower()
        if 'reward' not in table_text:
            continue

        reward_cells = reward_table.select('tr:nth-child(2) td')
        for cell in reward_cells:
            reward_link = cell.find('a', class_='a-link')
            if reward_link:
                reward_name = reward_link.text.strip()
                last_content = cell.contents[-1] if cell.contents else None
                reward_count = last_content.strip() if isinstance(last_content, str) else ''
                reward_count = reward_count.strip('×')
                if reward_count:  # Only add rewards with a count
                    rewards.append(f"{reward_name}: {reward_count}")

    return rewards

if __name__ == "__main__":
    base_url = "https://game8.co/games/Genshin-Impact/archives/297433"
    container_selector = "body > div.l-content > div.l-3col > div.l-3colMain > div.l-3colMain__center.l-3colMain__center--shadow > div.archive-style-wrapper"
    
    scrape_website(base_url, container_selector)
    print("Scraping completed.")